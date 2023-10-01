import json

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from baseClasses.BaseCharacter import BaseCharacter, getStatComments
from baseClasses.BaseEffect import BaseEffect
from estimator.DefaultEstimator import VisualizationInfo

from visualizer.visualizer import COLOR_DICT
import matplotlib.colors as mcolors

hex_dict = {
        'wind': '2ecc71',
        'fire': 'c75d5d',
        'ice': '3498db',
        'lightning': 'bc81df',
        'physical': 'bdc3c7',
        'quantum': '767bd4',
        'imaginary': 'dbb25a',
    }

DEBUG_COLUMN_NAMES = [
    'count',
    'ability',
    'damage',
    'energy',
    'skill points',
    'gauge',
    'action value',
    'SPD', 'ATK', 'HP', 'DEF',
    'DMG', 'CR', 'CD',
    'Vulnerability',
    'ResPen', 'DefShred',
]

def writeVisualizationList(visInfoList:list,path:str,sheetname:str):
    for i, visInfo in enumerate(visInfoList):
        if isinstance(visInfo,VisualizationInfo):
            visInfoList[i] = [visInfo]
    
    workbook:Workbook = Workbook()
    sheet:Worksheet = workbook.active
    sheet.title = sheetname

    current_row = 1
    left_align = Alignment(horizontal='left')
    decimal_format1 = NamedStyle(name='decimal_format1')
    decimal_format1.number_format = '0.0'
    decimal_format2 = NamedStyle(name='decimal_format2')
    decimal_format2.number_format = '0.00'
    decimal_format3 = NamedStyle(name='decimal_format3')
    decimal_format3.number_format = '0.000'
    
    # set column widths
    sheet.column_dimensions[get_column_letter(1)].width = 30
    sheet.column_dimensions[get_column_letter(2)].width = 60
    sheet.column_dimensions[get_column_letter(5)].width = 30
    
    for visInfo in visInfoList:
        
        rotationName:str = ' '.join([info.name for info in visInfo]) # append the names of the characters together
        character:BaseCharacter = visInfo[0].character # get the first character as the lead character
        totalEffect:BaseEffect = BaseEffect()
        for i, info in enumerate(visInfo):
            info:VisualizationInfo
            if i > 0: # only track energy and action values for the lead character
                info.effect.energy = 0.0
                info.effect.actionvalue = 0.0
                info.dotEffect.energy = 0.0
                info.dotEffect.actionvalue = 0.0
                info.breakEffect.energy = 0.0
                info.breakEffect.actionvalue = 0.0
            
            totalEffect += info.effect + info.breakEffect
            if sum(info.dotEffect.debugCount) > 0:
                totalEffect += info.dotEffect
        
        # write character info and header info
        color = hex_dict[character.element]
        current_cell = sheet.cell(row=current_row,column=1,value=rotationName)
        current_cell.font = Font(bold=True)
        
        for i, column_name in enumerate(DEBUG_COLUMN_NAMES):
            current_cell = sheet.cell(row=current_row, column=4+i, value=column_name)
            current_cell.font = Font(bold=True)
        
        initial_row = current_row
        current_row += 1
            
        # fill out character equipment
        equipment = [('Light Cone',character.lightcone.name),
                     ('Relic Set One',character.relicsetone.shortname),
                     ('Relic Set Two',character.relicsettwo.shortname),
                     ('Planar Set',character.planarset.shortname),
                     ('Mainstats',json.dumps(character.relicstats.mainstats)),
                     ('Substats',json.dumps(character.relicstats.substats))]
        
        for i, entry in enumerate(equipment):
            name, value = entry
            current_cell = sheet.cell(row=current_row+i,column=1,value=name)
            current_cell = sheet.cell(row=current_row+i,column=2,value=value)
        
        speed = character.getTotalStat('SPD')
        cycles = totalEffect.actionvalue * 100.0 / speed
        current_cell = sheet.cell(row=current_row,column=5,value='Total per Cycle')
        current_cell = sheet.cell(row=current_row,column=6,value=totalEffect.damage / cycles).style = decimal_format1
        current_cell = sheet.cell(row=current_row,column=8,value=totalEffect.skillpoints / cycles).style = decimal_format2
        current_cell = sheet.cell(row=current_row,column=9,value=totalEffect.gauge / cycles).style = decimal_format1
        current_row += 1
        
        current_cell = sheet.cell(row=current_row,column=5,value='Total per Rotation')
        current_cell = sheet.cell(row=current_row,column=6,value=totalEffect.damage).style = decimal_format1
        current_cell = sheet.cell(row=current_row,column=7,value=totalEffect.energy).style = decimal_format1
        current_cell = sheet.cell(row=current_row,column=8,value=totalEffect.skillpoints).style = decimal_format2
        current_cell = sheet.cell(row=current_row,column=9,value=totalEffect.gauge).style = decimal_format1
        current_cell = sheet.cell(row=current_row,column=10,value=totalEffect.actionvalue).style = decimal_format3
        current_row += 1
        
        # fill out rotation details
        for i, info in enumerate(totalEffect.debugInfo):
            current_cell = sheet.cell(row=current_row+i,column=4,value=totalEffect.debugCount[i])
            current_cell.style = decimal_format2
            
            for j, entry in enumerate(info):
                if isinstance(entry,list):
                    current_cell = sheet.cell(row=current_row+i,column=5+j,value=entry[0])
                    current_cell.comment = Comment(entry[1],'')
                else:
                    current_cell = sheet.cell(row=current_row+i,column=5+j,value=entry)
                
                if isinstance(current_cell.value,float):
                    if current_cell.value < 10.0 and current_cell.value != 0.0:
                        current_cell.style = decimal_format3
                    else:
                        current_cell.style = decimal_format1
        
        current_row += max(len(totalEffect.debugInfo),len(equipment)-2)+1
        
        for i in range(current_row-initial_row-1):
            for j in range(20):
                current_cell = sheet.cell(row=current_row-i-2,column=j+1)
                current_cell.fill = PatternFill(start_color=color,end_color=color,fill_type='solid')
                current_cell.alignment = left_align
    
    workbook.save(path)
    workbook.close()